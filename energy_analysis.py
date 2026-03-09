#!/usr/bin/env python3
"""
Energy Consumption Analysis from Grafana Data
Transcribes handwritten measurements and creates comprehensive Excel analysis
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from statistics import mean, stdev
import os

# Cluster 1 Energy consumption (grafana picks) - CORRECTED
CLUSTER1_DATA = {
    'Python Cold': [2814, 2810, 2812, 2820, 2808],
    'Python Warm': [2799, 2799, 2800, 2804, 2800],
    'JS Cold': [2809, 2809, 2817, 2811, 2825],
    'JS Warm': [2803, 2802, 2803, 2807, 2812],
    'Java Cold': [2834, 2841, 2828, 2840, 2825], 
    'Java Warm': [2810, 2817, 2804, 2804, 2804],
}

# Cluster 2 Energy consumption (grafana picks) - CORRECTED
CLUSTER2_DATA = {
    'Python Cold Start': [2977, 2996, 2960, 2960, 2970],
    'Python Warm Start': [2866, 2894, 2907, 2904, 2918],
    'JS Cold Start': [2962, 2958, 3023, 2965, 2957],
    'JS Warm Start': [2875, 2893, 2897, 2906, 2915],
    'Java Cold Start': [2974, 2983, 3018, 2954, 2992],
    'Java Warm Start': [2913, 2918, 2893, 2893, 2898],
}

BASELINE_IDLE = 2800  # Joules - baseline idle consumption

def calculate_stats(values):
    """Calculate statistics for a list of values"""
    return {
        'mean': mean(values),
        'stdev': stdev(values) if len(values) > 1 else 0,
        'min': min(values),
        'max': max(values),
        'count': len(values)
    }

def create_measurements_sheet(wb, cluster_name, data):
    """Create sheet with raw measurements"""
    ws = wb.create_sheet(f"{cluster_name} - Raw Data")
    
    # Title
    ws.append([f"{cluster_name} - Energy Consumption Measurements (Joules)"])
    ws.append([])
    
    # Headers
    headers = ['Test Type'] + [f'Iter {i+1}' for i in range(5)] + ['Mean', 'StdDev', 'Min', 'Max']
    ws.append(headers)
    
    # Data rows
    for test_name, values in data.items():
        stats = calculate_stats(values)
        row = [test_name] + values + [
            round(stats['mean'], 2),
            round(stats['stdev'], 2),
            stats['min'],
            stats['max']
        ]
        ws.append(row)
    
    # Styling
    style_header_row(ws, 3)
    auto_adjust_columns(ws)
    
    return ws

def create_overhead_analysis_sheet(wb, cluster_name, data):
    """Create sheet analyzing overhead vs baseline"""
    ws = wb.create_sheet(f"{cluster_name} - Overhead")
    
    # Title
    ws.append([f"{cluster_name} - Energy Overhead Analysis"])
    ws.append([f"Baseline Idle: {BASELINE_IDLE} J"])
    ws.append([])
    
    # Headers
    headers = ['Test Type', 'Mean (J)', 'Overhead (J)', 'Overhead (%)', 'StdDev (J)']
    ws.append(headers)
    
    # Calculate overhead for each test
    overhead_data = []
    for test_name, values in data.items():
        stats = calculate_stats(values)
        overhead_j = stats['mean'] - BASELINE_IDLE
        overhead_pct = (overhead_j / BASELINE_IDLE) * 100
        
        row = [
            test_name,
            round(stats['mean'], 2),
            round(overhead_j, 2),
            round(overhead_pct, 3),
            round(stats['stdev'], 2)
        ]
        ws.append(row)
        overhead_data.append({
            'test': test_name,
            'overhead_j': overhead_j,
            'overhead_pct': overhead_pct
        })
    
    # Styling
    style_header_row(ws, 4)
    auto_adjust_columns(ws)
    
    return ws, overhead_data

def create_cold_vs_hot_sheet(wb, cluster_name, data):
    """Create sheet comparing cold vs hot starts"""
    ws = wb.create_sheet(f"{cluster_name} - Cold vs Hot")
    
    # Title
    ws.append([f"{cluster_name} - Cold Start vs Warm Start Comparison"])
    ws.append([])
    
    # Headers
    headers = ['Language', 'Cold Mean (J)', 'Hot Mean (J)', 'Savings (J)', 'Savings (%)', 'Cold StdDev', 'Hot StdDev']
    ws.append(headers)
    
    # Find cold/hot pairs
    languages = set()
    for test_name in data.keys():
        lang = test_name.split()[0]  # Python, JS, Java
        languages.add(lang)
    
    for lang in sorted(languages):
        # Find cold and hot tests for this language
        cold_key = None
        hot_key = None
        
        for test_name in data.keys():
            if lang in test_name:
                if 'Cold' in test_name or 'cold' in test_name:
                    cold_key = test_name
                elif 'Warm' in test_name or 'warm' in test_name or 'Hot' in test_name:
                    hot_key = test_name
        
        if cold_key and hot_key:
            cold_stats = calculate_stats(data[cold_key])
            hot_stats = calculate_stats(data[hot_key])
            
            savings_j = cold_stats['mean'] - hot_stats['mean']
            savings_pct = (savings_j / cold_stats['mean']) * 100
            
            row = [
                lang,
                round(cold_stats['mean'], 2),
                round(hot_stats['mean'], 2),
                round(savings_j, 2),
                round(savings_pct, 2),
                round(cold_stats['stdev'], 2),
                round(hot_stats['stdev'], 2)
            ]
            ws.append(row)
    
    # Styling
    style_header_row(ws, 3)
    auto_adjust_columns(ws)
    
    return ws

def create_cluster_comparison_sheet(wb, c1_data, c2_data):
    """Compare energy consumption between clusters"""
    ws = wb.create_sheet("Cluster Comparison")
    
    # Title
    ws.append(["Cluster 1 (VM+Container) vs Cluster 2 (MicroVM) - Energy Comparison"])
    ws.append([])
    
    # Headers
    headers = ['Test Type', 'C1 Mean (J)', 'C2 Mean (J)', 'Difference (J)', 'Difference (%)', 'C1→C2 Impact']
    ws.append(headers)
    
    # Match tests between clusters
    test_mapping = {
        'Python Cold': 'Python Cold Start',
        'Python Warm': 'Python Warm Start',
        'JS Cold': 'JS Cold Start',
        'JS Warm': 'JS Warm Start',
        'Java Cold': 'Java Cold Start',
        'Java Warm': 'Java Warm Start'
    }
    
    for c1_test, c2_test in test_mapping.items():
        if c1_test in c1_data and c2_test in c2_data:
            c1_mean = mean(c1_data[c1_test])
            c2_mean = mean(c2_data[c2_test])
            diff = c2_mean - c1_mean
            diff_pct = (diff / c1_mean) * 100
            
            impact = "Higher" if diff > 0 else "Lower" if diff < 0 else "Same"
            
            row = [
                c1_test,
                round(c1_mean, 2),
                round(c2_mean, 2),
                round(diff, 2),
                round(diff_pct, 2),
                impact
            ]
            ws.append(row)
    
    # Styling
    style_header_row(ws, 3)
    auto_adjust_columns(ws)
    
    return ws

def create_summary_sheet(wb, c1_data, c2_data):
    """Create executive summary"""
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    
    # Title
    title = ws['A1']
    title.value = "Energy Consumption Analysis - Executive Summary"
    title.font = Font(size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Baseline
    ws[f'A{row}'] = "Baseline Idle Consumption:"
    ws[f'B{row}'] = f"{BASELINE_IDLE} J"
    ws[f'A{row}'].font = Font(bold=True)
    row += 2
    
    # Cluster 1 Summary
    ws[f'A{row}'] = "CLUSTER 1 (VM + Container)"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    for test_name, values in CLUSTER1_DATA.items():
        stats = calculate_stats(values)
        overhead = stats['mean'] - BASELINE_IDLE
        ws[f'A{row}'] = test_name
        ws[f'B{row}'] = f"{stats['mean']:.1f} J"
        ws[f'C{row}'] = f"+{overhead:.1f} J"
        ws[f'D{row}'] = f"+{(overhead/BASELINE_IDLE*100):.2f}%"
        row += 1
    
    row += 1
    
    # Cluster 2 Summary
    ws[f'A{row}'] = "CLUSTER 2 (MicroVM)"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    for test_name, values in CLUSTER2_DATA.items():
        stats = calculate_stats(values)
        overhead = stats['mean'] - BASELINE_IDLE
        ws[f'A{row}'] = test_name
        ws[f'B{row}'] = f"{stats['mean']:.1f} J"
        ws[f'C{row}'] = f"+{overhead:.1f} J"
        ws[f'D{row}'] = f"+{(overhead/BASELINE_IDLE*100):.2f}%"
        row += 1
    
    row += 2
    
    # Key Findings
    ws[f'A{row}'] = "KEY FINDINGS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    # Calculate averages
    c1_avg = mean([mean(vals) for vals in CLUSTER1_DATA.values()])
    c2_avg = mean([mean(vals) for vals in CLUSTER2_DATA.values()])
    
    ws[f'A{row}'] = f"• Cluster 1 Average: {c1_avg:.1f} J"
    row += 1
    ws[f'A{row}'] = f"• Cluster 2 Average: {c2_avg:.1f} J"
    row += 1
    ws[f'A{row}'] = f"• MicroVM Overhead: {c2_avg - c1_avg:+.1f} J ({((c2_avg - c1_avg)/c1_avg*100):+.2f}%)"
    row += 1
    
    # Set column widths manually
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Move to first position
    wb.move_sheet(ws, -(len(wb.sheetnames) - 1))
    
    return ws

def style_header_row(ws, row_num):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[row_num]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for col_num in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        
        for row in range(1, ws.max_row + 1):
            cell = ws[f'{column_letter}{row}']
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def main():
    print("\n" + "="*80)
    print("ENERGY CONSUMPTION ANALYSIS")
    print("="*80)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Cluster 1 sheets
    print("\nProcessing Cluster 1 (VM+Container)...")
    create_measurements_sheet(wb, "Cluster 1", CLUSTER1_DATA)
    create_overhead_analysis_sheet(wb, "Cluster 1", CLUSTER1_DATA)
    create_cold_vs_hot_sheet(wb, "Cluster 1", CLUSTER1_DATA)
    
    # Cluster 2 sheets
    print("Processing Cluster 2 (MicroVM)...")
    create_measurements_sheet(wb, "Cluster 2", CLUSTER2_DATA)
    create_overhead_analysis_sheet(wb, "Cluster 2", CLUSTER2_DATA)
    create_cold_vs_hot_sheet(wb, "Cluster 2", CLUSTER2_DATA)
    
    # Comparison sheets
    print("Creating comparison analysis...")
    create_cluster_comparison_sheet(wb, CLUSTER1_DATA, CLUSTER2_DATA)
    
    # Summary (will be moved to first position)
    create_summary_sheet(wb, CLUSTER1_DATA, CLUSTER2_DATA)
    
    # Save
    output_file = "Energy_Consumption_Analysis.xlsx"
    wb.save(output_file)
    
    print(f"\n✅ Analysis complete!")
    print(f"   File saved: {output_file}")
    print(f"   Total sheets: {len(wb.sheetnames)}")
    print("\nSheets created:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
    
    # Print summary statistics
    print("\n" + "="*80)
    print("SUMMARY STATISTICS")
    print("="*80)
    
    c1_avg = mean([mean(vals) for vals in CLUSTER1_DATA.values()])
    c2_avg = mean([mean(vals) for vals in CLUSTER2_DATA.values()])
    
    print(f"\nCluster 1 (VM+Container):")
    print(f"  Average consumption: {c1_avg:.1f} J")
    print(f"  Average overhead: {c1_avg - BASELINE_IDLE:.1f} J (+{((c1_avg - BASELINE_IDLE)/BASELINE_IDLE*100):.2f}%)")
    
    print(f"\nCluster 2 (MicroVM):")
    print(f"  Average consumption: {c2_avg:.1f} J")
    print(f"  Average overhead: {c2_avg - BASELINE_IDLE:.1f} J (+{((c2_avg - BASELINE_IDLE)/BASELINE_IDLE*100):.2f}%)")
    
    print(f"\nMicroVM Impact:")
    print(f"  Difference: {c2_avg - c1_avg:+.1f} J")
    print(f"  Percentage: {((c2_avg - c1_avg)/c1_avg*100):+.2f}%")
    
    # Detailed breakdown
    print(f"\n" + "-"*80)
    print("COLD vs WARM SAVINGS:")
    print("-"*80)
    
    # Cluster 1
    print("\nCluster 1:")
    for lang in ['Python', 'JS', 'Java']:
        cold_key = f'{lang} Cold'
        warm_key = f'{lang} Warm'
        if cold_key in CLUSTER1_DATA and warm_key in CLUSTER1_DATA:
            cold_mean = mean(CLUSTER1_DATA[cold_key])
            warm_mean = mean(CLUSTER1_DATA[warm_key])
            savings = cold_mean - warm_mean
            savings_pct = (savings / cold_mean) * 100
            print(f"  {lang:10}: Cold={cold_mean:.1f}J, Warm={warm_mean:.1f}J, Savings={savings:.1f}J ({savings_pct:.1f}%)")
    
    # Cluster 2
    print("\nCluster 2:")
    for lang in ['Python', 'JS', 'Java']:
        cold_key = f'{lang} Cold Start'
        warm_key = f'{lang} Warm Start'
        if cold_key in CLUSTER2_DATA and warm_key in CLUSTER2_DATA:
            cold_mean = mean(CLUSTER2_DATA[cold_key])
            warm_mean = mean(CLUSTER2_DATA[warm_key])
            savings = cold_mean - warm_mean
            savings_pct = (savings / cold_mean) * 100
            print(f"  {lang:10}: Cold={cold_mean:.1f}J, Warm={warm_mean:.1f}J, Savings={savings:.1f}J ({savings_pct:.1f}%)")
    
    print("\n" + "="*80)

if __name__ == '__main__':
    main()
