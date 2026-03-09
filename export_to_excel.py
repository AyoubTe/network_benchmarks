#!/usr/bin/env python3
"""
Export benchmark results to Excel with multiple sheets
One sheet per test type with all iterations
"""

import json
import os
import glob
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def load_all_results(base_dir):
    """Load all JSON results and organize by test type"""
    results = defaultdict(list)
    
    # Find all JSON files recursively
    json_files = glob.glob(os.path.join(base_dir, '**', '*.json'), recursive=True)
    
    for json_file in json_files:
        try:
            with open(json_file, 'r') as f:
                data = json.load(f)
                
                # Determine test key from data
                if 'language' in data and 'cold_start' in data:
                    lang = data['language']
                    start_type = 'cold' if data['cold_start'] else 'hot'
                    iteration = data.get('iteration', 0)
                    
                    key = f"{lang}_{start_type}"
                    results[key].append({
                        'file': os.path.basename(json_file),
                        'data': data,
                        'iteration': iteration
                    })
        except Exception as e:
            print(f"Error loading {json_file}: {e}")
    
    # Sort by iteration number
    for key in results:
        results[key].sort(key=lambda x: x['iteration'])
    
    return results

def create_summary_sheet(results):
    """Create overview summary sheet"""
    summary_data = []
    
    for test_key in sorted(results.keys()):
        test_results = results[test_key]
        
        # Extract invocation times
        inv_times = [r['data'].get('invocation_time_s', 0) for r in test_results]
        
        # Get network metrics from first iteration
        if test_results and 'benchmarks' in test_results[0]['data']:
            benchmarks = test_results[0]['data']['benchmarks']
            tcp = benchmarks.get('tcp_latency', {}).get('avg_ms', 0)
            http = benchmarks.get('http_latency', {}).get('avg_ms', 0)
            dns = benchmarks.get('dns_resolution', {}).get('avg_ms', 0)
            bw = benchmarks.get('bandwidth', {}).get('bandwidth_mbps', 0)
        else:
            tcp = http = dns = bw = 0
        
        summary_data.append({
            'Test': test_key,
            'Iterations': len(inv_times),
            'Mean Time (s)': sum(inv_times) / len(inv_times) if inv_times else 0,
            'Min Time (s)': min(inv_times) if inv_times else 0,
            'Max Time (s)': max(inv_times) if inv_times else 0,
            'TCP Latency (ms)': tcp,
            'HTTP Latency (ms)': http,
            'DNS Latency (ms)': dns,
            'Bandwidth (Mbps)': bw
        })
    
    return pd.DataFrame(summary_data)

def create_test_detail_sheet(test_results):
    """Create detailed sheet for one test with all iterations"""
    rows = []
    
    for result in test_results:
        data = result['data']
        iteration = result['iteration']
        
        # Basic info
        row = {
            'Iteration': iteration,
            'Invocation Time (s)': data.get('invocation_time_s', 0),
            'Cold Start': data.get('cold_start', False),
            'Timestamp': data.get('timestamp', 0),
            'Hostname': data.get('hostname', ''),
            'Pod Name': data.get('pod_name', 'N/A')
        }
        
        # Network metrics
        if 'benchmarks' in data:
            benchmarks = data['benchmarks']
            
            # TCP Latency
            if 'tcp_latency' in benchmarks:
                tcp = benchmarks['tcp_latency']
                row['TCP Avg (ms)'] = tcp.get('avg_ms', 0)
                row['TCP Min (ms)'] = tcp.get('min_ms', 0)
                row['TCP Max (ms)'] = tcp.get('max_ms', 0)
                row['TCP StdDev (ms)'] = tcp.get('stdev_ms', 0)
            
            # HTTP Latency
            if 'http_latency' in benchmarks:
                http = benchmarks['http_latency']
                row['HTTP Avg (ms)'] = http.get('avg_ms', 0)
                row['HTTP Min (ms)'] = http.get('min_ms', 0)
                row['HTTP Max (ms)'] = http.get('max_ms', 0)
                row['HTTP StdDev (ms)'] = http.get('stdev_ms', 0)
                row['HTTP Median (ms)'] = http.get('median_ms', 0)
            
            # DNS Resolution
            if 'dns_resolution' in benchmarks:
                dns = benchmarks['dns_resolution']
                row['DNS Avg (ms)'] = dns.get('avg_ms', 0)
                row['DNS Min (ms)'] = dns.get('min_ms', 0)
                row['DNS Max (ms)'] = dns.get('max_ms', 0)
            
            # Bandwidth
            if 'bandwidth' in benchmarks:
                bw = benchmarks['bandwidth']
                row['BW Mbps'] = bw.get('bandwidth_mbps', 0)
                row['BW Avg Download (s)'] = bw.get('avg_download_time_s', 0)
                row['BW File Size (MB)'] = bw.get('file_size_mb', 0)
                row['BW Min Time (s)'] = bw.get('min_time_s', 0)
                row['BW Max Time (s)'] = bw.get('max_time_s', 0)
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_network_details_sheet(test_results):
    """Create sheet with detailed network latency measurements per iteration"""
    rows = []
    
    for result in test_results:
        data = result['data']
        iteration = result['iteration']
        
        if 'benchmarks' not in data:
            continue
        
        benchmarks = data['benchmarks']
        
        # TCP latencies
        if 'tcp_latency' in benchmarks and 'latencies_ms' in benchmarks['tcp_latency']:
            for i, latency in enumerate(benchmarks['tcp_latency']['latencies_ms'], 1):
                rows.append({
                    'Iteration': iteration,
                    'Metric': 'TCP',
                    'Measurement': i,
                    'Latency (ms)': latency
                })
        
        # HTTP latencies
        if 'http_latency' in benchmarks and 'latencies_ms' in benchmarks['http_latency']:
            for i, latency in enumerate(benchmarks['http_latency']['latencies_ms'], 1):
                rows.append({
                    'Iteration': iteration,
                    'Metric': 'HTTP',
                    'Measurement': i,
                    'Latency (ms)': latency
                })
        
        # Bandwidth download times
        if 'bandwidth' in benchmarks and 'download_times_s' in benchmarks['bandwidth']:
            for i, time in enumerate(benchmarks['bandwidth']['download_times_s'], 1):
                rows.append({
                    'Iteration': iteration,
                    'Metric': 'Bandwidth',
                    'Measurement': i,
                    'Time (s)': time
                })
    
    return pd.DataFrame(rows)

def style_worksheet(ws, title):
    """Apply styling to worksheet"""
    # Title row
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Header row styling (row 3 since we have title and blank row)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling to row 3
    for cell in ws[3]:
        if cell.value:  # Only style cells with values
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def export_to_excel(cluster_name, base_dir, output_file):
    """Main export function"""
    print(f"\nExporting {cluster_name} results to Excel...")
    print(f"Reading from: {base_dir}")
    
    # Load all results
    results = load_all_results(base_dir)
    
    if not results:
        print(f"❌ No results found in {base_dir}")
        return
    
    print(f"Found {len(results)} test types:")
    for key in sorted(results.keys()):
        print(f"  - {key}: {len(results[key])} iterations")
    
    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Add summary sheet
    print("\nCreating summary sheet...")
    summary_df = create_summary_sheet(results)
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append([f"{cluster_name} - Results Summary"])
    ws_summary.append([])
    
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(r)
    
    style_worksheet(ws_summary, f"{cluster_name} - Summary")
    
    # Add detailed sheet for each test
    for test_key in sorted(results.keys()):
        print(f"Creating sheet for {test_key}...")
        
        # Main details sheet
        test_df = create_test_detail_sheet(results[test_key])
        ws_test = wb.create_sheet(f"{test_key}")
        ws_test.append([f"{cluster_name} - {test_key.upper()}"])
        ws_test.append([])
        
        for r in dataframe_to_rows(test_df, index=False, header=True):
            ws_test.append(r)
        
        style_worksheet(ws_test, f"{cluster_name} - {test_key}")
        
        # Network details sheet
        network_df = create_network_details_sheet(results[test_key])
        if not network_df.empty:
            ws_network = wb.create_sheet(f"{test_key}_net")
            ws_network.append([f"{cluster_name} - {test_key.upper()} - Network Details"])
            ws_network.append([])
            
            for r in dataframe_to_rows(network_df, index=False, header=True):
                ws_network.append(r)
            
            style_worksheet(ws_network, f"{cluster_name} - {test_key} - Network")
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✅ Excel file saved: {output_file}")
    print(f"   Total sheets: {len(wb.sheetnames)}")

if __name__ == '__main__':
    # Export Cluster 1
    export_to_excel(
        "Cluster 1 (VM+Container)",
        "./cluster_1_results",
        "./Cluster1_Benchmark_Results.xlsx"
    )
    
    # Export Cluster 2
    export_to_excel(
        "Cluster 2 (MicroVM)",
        "./cluster_2_results",
        "./Cluster2_Benchmark_Results.xlsx"
    )
    
    print("\n" + "="*80)
    print("EXPORT COMPLETE!")
    print("="*80)
    print("\nGenerated files:")
    print("  1. Cluster1_Benchmark_Results.xlsx")
    print("  2. Cluster2_Benchmark_Results.xlsx")
    print("\nEach file contains:")
    print("  - Summary sheet with all tests")
    print("  - Detailed sheet per test (python_cold, python_hot, etc.)")
    print("  - Network details sheet per test with individual measurements")
    print("="*80)
